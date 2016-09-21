"""Hopelessly late 'unit tests' for the reinserter."""
# At this point I need to test various things about the data, not really the program itself.
# So this doesn't really count much for 'test coverage.' Oh well.

# For this case, unit tests are not useful - I need to check things about the data itself.
# Unit tests stop after one thing is wrong; I need a list of all things that are wrong.

from utils import DUMP_XLS, onscreen_length
from openpyxl import load_workbook
from rominfo import file_blocks, CREATURE_BLOCK, CREATURE_MAX_LENGTH, DIALOGUE_MAX_LENGTH, DAT_MAX_LENGTH, FULLSCREEN_MAX_LENGTH

def test_increasing_offsets():
    """Make sure the offsets are strictly increasing - so that no strings are mislabeled."""
    wb = load_workbook(DUMP_XLS)
    sheets = wb.get_sheet_names()
    sheets.remove('ORIGINAL')
    sheets.remove('MISC TITLES')
    sheets.remove('SINKA.DAT')
    sheets.remove('SEND.DAT')
    for sheet in sheets:
        ws = wb.get_sheet_by_name(sheet)
        new_offset = 0
        for index, row in enumerate(ws.rows[1:]):  # Skip the first row, it's just labels
            try:
                this_offset = int(row[0].value, 16)
            except TypeError:
                break
            if this_offset <= new_offset:
                print "In sheet %s, fix offset at row %s; %s < %s" % (sheet, index+2, hex(int(row[0].value, 16)), hex(new_offset))
            new_offset = this_offset

def test_substrings_of_earlier_strings():
    wb = load_workbook(DUMP_XLS)
    sheets = wb.get_sheet_names()
    sheets.remove('ORIGINAL')
    sheets.remove('MISC TITLES')
    sheets.remove('SINKA.DAT')
    sheets.remove('SEND.DAT')
    for gamefile in file_blocks:
        ws = wb.get_sheet_by_name(gamefile)
        for (start, stop) in file_blocks[gamefile]:
            previous_untranslated_jp = []
            for index, row in enumerate(ws.rows[1:]):
                try:
                    this_offset = int(row[0].value, 16)
                except TypeError:
                    break
                if start <= this_offset <= stop:
                    jp_string = row[2].value 
                    if isinstance(jp_string, basestring):
                        for i in previous_untranslated_jp:    # Repeats (nametags, etc) are also substrings!
                            if jp_string in i[0] and row[4].value:     # If it's occurred before AND it's been translated:
                                print "%s, '%s' is a substring of the untranslated string at %s" % (gamefile, row[4].value, i[1])
                                # More informative way of doing this...? Terminal can't display jp text though.
                                break
                        # Add it to the 
                        if row[4].value is None:
                            previous_untranslated_jp.append((jp_string, row[0].value))

def test_all_string_lengths():
    """Corasest string test. No string can be longer than 77."""
    wb = load_workbook(DUMP_XLS)
    sheets = wb.get_sheet_names()
    sheets.remove('ORIGINAL')
    sheets.remove('MISC TITLES')
    for sheet in sheets:
        ws = wb.get_sheet_by_name(sheet)
        new_offset = 0
        for index, row in enumerate(ws.rows[1:]):
            if row[4].value:
                print row[4].value
                if onscreen_length(row[4].value) > FULLSCREEN_MAX_LENGTH:
                    print "%s %s: %s has length %s" % (sheet, row[0].value, row[4].value, onscreen_length(row[4].value))

def test_creature_string_lengths():
    """No creature name can have a name longer than 21."""
    wb = load_workbook(DUMP_XLS)
    for sheet in ['ST1.EXE', 'ST2.EXE', 'ST3.EXE', 'ST4.EXE', 'ST5.EXE', 'ST6.EXE']:
        creature_lo, creature_hi = CREATURE_BLOCK[sheet]
        ws = wb.get_sheet_by_name(sheet)
        new_offset = 0
        for index, row in enumerate(ws.rows[1:]):
            if int(row[0].value, 16) >= creature_lo and int(row[0].value, 16) <= creature_hi:
                if row[4].value:
                    assert onscreen_length(row[4].value) <= CREATURE_MAX_LENGTH, "In sheet %s, shorten string at row %s" % (sheet, index+2)

def test_dat_string_lengths():
    """No encyclopedia string should be longer than 68."""
    wb = load_workbook(DUMP_XLS)
    for sheet in ['SEND.DAT', 'SINKA.DAT']:
        ws = wb.get_sheet_by_name(sheet)
        new_offset = 0
        for index, row in enumerate(ws.rows[1:]):
            if row[4].value:
                assert onscreen_length(row[4].value) <= DAT_MAX_LENGTH, "In sheet %s, shorten string at row %s" % (sheet, index+2)

def test_game_string_lengths():
    """
    Wide onscreen strings can't be more than 68? characters.
    Ones not marked as wide can't be more than 42 characters.
    """
    wb = load_workbook(DUMP_XLS)
    for sheet in ['ST1.EXE', 'ST2.EXE', 'ST3.EXE', 'ST4.EXE', 'ST5.EXE', 'ST5S1.EXE',
                  'ST5S2.EXE', 'ST5S3.EXE', "ST6.EXE"]:
        ws = wb.get_sheet_by_name(sheet)
        new_offset = 0
        for index, row in enumerate(ws.rows[1:]):
            if row[7].value:
                if row[7].value == 'wide':
                    if onscreen_length(row[4].value) > FULLSCREEN_MAX_LENGTH:
                        try:
                            print "%s %s: %s has length %s" % (sheet, row[0].value, row[4].value, onscreen_length(row[4].value))
                        except UnicodeDecodeError:
                            print "%s %s: %s has length %s" % (sheet, row[0].value, 'some string', onscreen_length(row[4].value))
            else:
                if onscreen_length(row[4].value) > DIALOGUE_MAX_LENGTH:
                    try:
                        print "%s %s: %s has length %s" % (sheet, row[0].value, row[4].value, onscreen_length(row[4].value))
                    except UnicodeDecodeError:
                        print "%s %s: %s has length %s" % (sheet, row[0].value, 'some string', onscreen_length(row[4].value))

# make sure a blank translation sheet doesn't return overflow errors

if __name__ == '__main__':
    test_increasing_offsets()
    test_game_string_lengths()
    test_substrings_of_earlier_strings()
